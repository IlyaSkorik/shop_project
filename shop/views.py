from io import BytesIO

from django.conf import settings
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import ensure_csrf_cookie
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework import permissions, viewsets
from rest_framework.response import Response

from .models import Product, Category, Manufacturer, Cart, CartItem, Order, OrderItem
from .serializers import (
    CartItemSerializer,
    CartSerializer,
    CategorySerializer,
    ManufacturerSerializer,
    OrderItemSerializer,
    OrderSerializer,
    ProductSerializer,
)

@ensure_csrf_cookie
def home_view(request):
    popular_products = Product.objects.select_related("category", "manufacturer").order_by("-id")[:6]
    categories = Category.objects.all()
    return render(request, 'shop/home.html', {
        'popular_products': popular_products,
        'categories': categories,
    })

def author_view(request):
    return HttpResponse("Сайт разработал: Скорик Илья, группа 87ТП")

def about_view(request):
    return HttpResponse('Добро пожаловать в хобби-гипермаркет "Леонард"')


@ensure_csrf_cookie
def product_list(request):
    products = Product.objects.select_related("category", "manufacturer").order_by("name")
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()

    category_id = request.GET.get('category')
    manufacturer_id = request.GET.get('manufacturer')
    search_query = request.GET.get('search')

    if category_id:
        products = products.filter(category_id=category_id)

    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    paginator = Paginator(products, 9)
    page_obj = paginator.get_page(request.GET.get('page'))

    query_params = request.GET.copy()
    query_params.pop('page', None)

    context = {
        'products': page_obj,
        'page_obj': page_obj,
        'categories': categories,
        'manufacturers': manufacturers,
        'selected_category': category_id or '',
        'selected_manufacturer': manufacturer_id or '',
        'search_query': search_query or '',
        'query_string': query_params.urlencode(),
    }

    return render(request, 'shop/product_list.html', context)

@ensure_csrf_cookie
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'shop/product_detail.html', {'product': product})

@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    cart, created = Cart.objects.get_or_create(user=request.user)

    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={'quantity': 1} 
    )

    if not created:
        if cart_item.quantity + 1 <= product.stock_quantity:
            cart_item.quantity += 1
            cart_item.save()
        else:
            messages.error(request, "Недостаточно товара на складе.")

    return redirect('cart')

@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)

    if request.method == 'POST':
        quantity = int(request.POST.get('quantity'))

        if quantity <= cart_item.product.stock_quantity:
            cart_item.quantity = quantity
            cart_item.save()
        else:
            messages.error(request, "Количество превышает остаток на складе.")

    return redirect('cart')

@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()
    return redirect('cart')

@login_required
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    items = cart.items.all()
    total_price = cart.total_price()

    return render(request, 'shop/cart.html', {
        'cart': cart,
        'items': items,
        'total_price': total_price
    })


def build_receipt(order):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Receipt"

    sheet["A1"] = f"Чек по заказу #{order.id}"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A3"] = "Покупатель"
    sheet["B3"] = order.user.username
    sheet["A4"] = "Email"
    sheet["B4"] = order.email
    sheet["A5"] = "Адрес доставки"
    sheet["B5"] = order.shipping_address
    sheet["A6"] = "Дата заказа"
    sheet["B6"] = order.created_at.strftime("%d.%m.%Y %H:%M")

    headers = ["Товар", "Количество", "Цена", "Сумма"]
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[8]:
        cell.font = Font(bold=True)

    for item in order.items.all():
        sheet.append([
            item.product_name,
            item.quantity,
            float(item.price),
            float(item.total_price()),
        ])

    total_row = sheet.max_row + 1
    sheet.cell(total_row, 3, "Итого")
    sheet.cell(total_row, 4, float(order.total_price))
    sheet.cell(total_row, 3).font = Font(bold=True)
    sheet.cell(total_row, 4).font = Font(bold=True)

    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 15

    receipt = BytesIO()
    workbook.save(receipt)
    receipt.seek(0)
    return receipt


def send_order_receipt(order):
    receipt = build_receipt(order)
    email = EmailMessage(
        subject=f"Чек по заказу #{order.id}",
        body=(
            f"Здравствуйте, {order.user.username}!\n\n"
            f"Ваш заказ #{order.id} оформлен.\n"
            f"Сумма заказа: {order.total_price} BYN.\n"
            "Чек прикреплен к письму."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[order.email],
    )
    email.attach(
        filename=f"receipt_order_{order.id}.xlsx",
        content=receipt.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    email.send()


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def api_add_to_cart(request):
    product_id = request.data.get("product_id")
    quantity = int(request.data.get("quantity", 1))

    if quantity < 1:
        return Response({"detail": "Количество должно быть больше нуля."}, status=status.HTTP_400_BAD_REQUEST)

    product = get_object_or_404(Product, id=product_id)
    if product.stock_quantity < 1:
        return Response({"detail": "Товара нет в наличии."}, status=status.HTTP_400_BAD_REQUEST)

    cart, _ = Cart.objects.get_or_create(user=request.user)
    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={"quantity": 0},
    )

    if cart_item.quantity + quantity > product.stock_quantity:
        return Response({"detail": "Недостаточно товара на складе."}, status=status.HTTP_400_BAD_REQUEST)

    cart_item.quantity += quantity
    cart_item.save()

    return Response({
        "detail": "Товар добавлен в корзину.",
        "item_id": cart_item.id,
        "quantity": cart_item.quantity,
        "created": created,
    })


@login_required
def checkout(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    items = list(cart.items.select_related("product"))

    if not items:
        messages.error(request, "Корзина пуста. Добавьте товары перед оформлением заказа.")
        return redirect("cart")

    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        shipping_address = request.POST.get("shipping_address", "").strip()

        if not email or not shipping_address:
            messages.error(request, "Введите email и адрес доставки.")
            return redirect("checkout")

        with transaction.atomic():
            total_price = cart.total_price()
            order = Order.objects.create(
                user=request.user,
                email=email,
                shipping_address=shipping_address,
                total_price=total_price,
            )

            for item in items:
                product = item.product
                if item.quantity > product.stock_quantity:
                    messages.error(
                        request,
                        f"Недостаточно товара '{product.name}' на складе.",
                    )
                    order.delete()
                    return redirect("cart")

                OrderItem.objects.create(
                    order=order,
                    product=product,
                    product_name=product.name,
                    price=product.price,
                    quantity=item.quantity,
                )
                product.stock_quantity -= item.quantity
                product.save(update_fields=["stock_quantity"])

            cart.items.all().delete()

        send_order_receipt(order)
        messages.success(request, "Заказ оформлен. Чек отправлен на email.")
        return redirect("product_list")

    return render(request, "shop/checkout.html", {
        "items": items,
        "total_price": cart.total_price(),
        "default_email": request.user.email,
    })


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    def get_queryset(self):
        queryset = Product.objects.select_related("category", "manufacturer").order_by("name")
        category_id = self.request.query_params.get("category")
        manufacturer_id = self.request.query_params.get("manufacturer")
        search_query = self.request.query_params.get("search")

        if category_id:
            queryset = queryset.filter(category_id=category_id)
        if manufacturer_id:
            queryset = queryset.filter(manufacturer_id=manufacturer_id)
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(description__icontains=search_query)
            )

        return queryset


class CartViewSet(viewsets.ModelViewSet):
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Cart.objects.prefetch_related("items__product").select_related("user")
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class CartItemViewSet(viewsets.ModelViewSet):
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = CartItem.objects.select_related("cart__user", "product")
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(cart__user=self.request.user)


class OrderViewSet(viewsets.ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Order.objects.prefetch_related("items__product").select_related("user")
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user, total_price=0)


class OrderItemViewSet(viewsets.ModelViewSet):
    serializer_class = OrderItemSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = OrderItem.objects.select_related("order__user", "product")
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(order__user=self.request.user)

    def perform_create(self, serializer):
        product = serializer.validated_data["product"]
        serializer.save(product_name=product.name, price=product.price)
